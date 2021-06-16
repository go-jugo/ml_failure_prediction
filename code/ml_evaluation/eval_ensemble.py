from sklearn.model_selection import cross_val_score, cross_validate
from sklearn.model_selection import cross_val_predict
from sklearn.model_selection import StratifiedShuffleSplit, StratifiedKFold, train_test_split
import sklearn
import statistics
import pandas as pd
import copy
from sklearn.preprocessing import MinMaxScaler
from sklearn.svm import LinearSVC
from sklearn import svm
import numpy as np
from collections import Counter
from imblearn.over_sampling import SMOTE, BorderlineSMOTE, SVMSMOTE
from imblearn.under_sampling import RandomUnderSampler
from imblearn.ensemble import BalancedBaggingClassifier
from monitoring.time_it import timing
from sklearn.metrics import make_scorer, accuracy_score, precision_score, recall_score, f1_score, confusion_matrix, classification_report
import xlsxwriter
from mlxtend.classifier import EnsembleVoteClassifier
from sklearn.svm import LinearSVC, SVC
from openpyxl import load_workbook
from types import SimpleNamespace
import ast
import glob
from math import ceil

def eval_ensemble(X, y, config, crossvalidation, ml_list, sampling_percentage, random_state, eval_each_clf=True):

    print('Configurations: ' + str(config))

    if len(config['random_state']) != len(config['ml_list']):
        raise ValueError('Number of classifiers and random states are not the same')

    def tp(y_true, y_pred): return confusion_matrix(y_true, y_pred, labels=[config['target_errorCode'], -1])[0, 0]
    def tn(y_true, y_pred): return confusion_matrix(y_true, y_pred, labels=[config['target_errorCode'], -1])[1, 1]
    def fp(y_true, y_pred): return confusion_matrix(y_true, y_pred, labels=[config['target_errorCode'], -1])[1, 0]
    def fn(y_true, y_pred): return confusion_matrix(y_true, y_pred, labels=[config['target_errorCode'], -1])[0, 1]

    scoring = {'accuracy': make_scorer(accuracy_score),
               'precision': make_scorer(precision_score, pos_label=config['target_errorCode']),
               'recall': make_scorer(recall_score, pos_label=config['target_errorCode']),
               'f1_score': make_scorer(f1_score, pos_label=config['target_errorCode']),
               'precision_neg': make_scorer(precision_score, pos_label=-1),
               'recall_neg': make_scorer(recall_score, pos_label=-1),
               'f1_score_neg': make_scorer(f1_score, pos_label=-1),
               'tp': make_scorer(tp),
               'tn': make_scorer(tn),
               'fp': make_scorer(fp),
               'fn': make_scorer(fn)}

    scores = {}
    for element in scoring:
        scores['test_' + element] = []

    cv = StratifiedKFold(n_splits=crossvalidation, shuffle=False, random_state=None)

    if eval_each_clf:
        all_scores = {}
        for i in range(len(random_state)):
            all_scores['clf_' + str(i)] = copy.deepcopy(scores)
            all_scores['clf_' + str(i)]['random_state'] = [random_state[i]]

    for train_idx, test_idx, in cv.split(X, y):
        X_train, y_train = X.iloc[train_idx], y.iloc[train_idx]
        X_test, y_test = X.iloc[test_idx], y.iloc[test_idx]
        rus_test = RandomUnderSampler(random_state=42)
        X_test_rus, y_test_rus = rus_test.fit_resample(X_test, y_test)
        clf_dict = {}
        counter = 0
        for state in random_state:
            rus = RandomUnderSampler(random_state=state)
            X_train_rus, y_train_rus = rus.fit_resample(X_train, y_train)
            clf_dict['clf_' + str(counter)] = ml_list[counter]
            clf_dict['clf_' + str(counter)].fit(X_train_rus, y_train_rus)
            counter += 1
        eclf = EnsembleVoteClassifier(clfs=[clf_dict[element] for element in clf_dict], voting=config['ensemble_voting'], refit=False)
        eclf.fit(None, np.array([config['target_errorCode'], -1]))
        y_pred = eclf.predict(X_test_rus)
        scores_dict = classification_report(y_test_rus, y_pred, output_dict=True)
        scores['test_accuracy'].append(scores_dict['accuracy'])
        scores['test_precision'].append(scores_dict[str(config['target_errorCode'])]['precision'])
        scores['test_precision_neg'].append(scores_dict['-1']['precision'])
        scores['test_recall'].append(scores_dict[str(config['target_errorCode'])]['recall'])
        scores['test_recall_neg'].append(scores_dict['-1']['recall'])
        scores['test_f1_score'].append(scores_dict[str(config['target_errorCode'])]['f1-score'])
        scores['test_f1_score_neg'].append(scores_dict['-1']['f1-score'])
        scores['test_tp'].append(tp(y_test_rus, y_pred))
        scores['test_tn'].append(tn(y_test_rus, y_pred))
        scores['test_fp'].append(fp(y_test_rus, y_pred))
        scores['test_fn'].append(fn(y_test_rus, y_pred))
        if eval_each_clf:
            for element in all_scores:
                y_pred = clf_dict[element].predict(X_test_rus)
                scores_dict = classification_report(y_test_rus, y_pred, output_dict=True)
                all_scores[element]['test_accuracy'].append(scores_dict['accuracy'])
                all_scores[element]['test_precision'].append(scores_dict[str(config['target_errorCode'])]['precision'])
                all_scores[element]['test_precision_neg'].append(scores_dict['-1']['precision'])
                all_scores[element]['test_recall'].append(scores_dict[str(config['target_errorCode'])]['recall'])
                all_scores[element]['test_recall_neg'].append(scores_dict['-1']['recall'])
                all_scores[element]['test_f1_score'].append(scores_dict[str(config['target_errorCode'])]['f1-score'])
                all_scores[element]['test_f1_score_neg'].append(scores_dict['-1']['f1-score'])
                all_scores[element]['test_tp'].append(tp(y_test_rus, y_pred))
                all_scores[element]['test_tn'].append(tn(y_test_rus, y_pred))
                all_scores[element]['test_fp'].append(fp(y_test_rus, y_pred))
                all_scores[element]['test_fn'].append(fn(y_test_rus, y_pred))
    for element in scores:
        scores[element] = np.array(scores[element])

    if eval_each_clf:
        for clf in all_scores:
            for element in all_scores[clf]:
                all_scores[clf][element] = np.array(all_scores[clf][element])

    for key, value in scores.items():
        print(str(key))
        print(value)
        print('M: ' + str(value.mean()))
        print('SD: ' + str(value.std()))



    wb = load_workbook(filename='../results/Evaluation_results_EUS.xlsx')
    ws1 = wb.active
    counter = len(list(ws1.rows))
    n = len(X)
    n_error_class = len(y[y == config['target_errorCode']])
    n_non_error_class = len(y[y != config['target_errorCode']])
    sampling_frequency = config['sampling_frequency']
    ts_fresh_window_length = config['ts_fresh_window_length']
    ts_fresh_window_end = config['ts_fresh_window_end']
    ts_fresh_minimal_features = config['ts_fresh_minimal_features']
    target_col = config['target_col']
    target_errorCode = config['target_errorCode']
    rand_state = str(random_state)
    sampling_percentage = config['sampling_percentage']
    balance = config['balance_ratio']
    n_classifier = len(config['ml_list'])
    ensemble_voting = config['ensemble_voting']
    ml_algorithm = str(config['ml_list'])
    cv = config['cv']
    Accuracy = str(scores['test_accuracy'])
    Accuracy_mean = scores['test_accuracy'].mean()
    Accuracy_std = scores['test_accuracy'].std()
    Precision = str(scores['test_precision'])
    Precision_neg = str(scores['test_precision_neg'])
    Precision_mean = scores['test_precision'].mean()
    Precision_mean_neg = scores['test_precision_neg'].mean()
    Precision_std = scores['test_precision'].std()
    Precision_std_neg = scores['test_precision_neg'].std()
    Recall = str(scores['test_recall'])
    Recall_neg = str(scores['test_recall_neg'])
    Recall_mean = scores['test_recall'].mean()
    Recall_mean_neg = scores['test_recall_neg'].mean()
    Recall_std = scores['test_recall'].std()
    Recall_std_neg = scores['test_recall_neg'].std()
    F1_Score = str(scores['test_f1_score'])
    F1_Score_neg = str(scores['test_f1_score_neg'])
    F1_Score_mean = scores['test_f1_score'].mean()
    F1_Score_mean_neg = scores['test_f1_score_neg'].mean()
    F1_Score_std = scores['test_f1_score'].std()
    F1_Score_std_neg = scores['test_f1_score_neg'].std()
    tp_cv = str(scores['test_tp'])
    tn_cv = str(scores['test_tn'])
    fp_cv = str(scores['test_fp'])
    fn_cv = str(scores['test_fn'])
    tp_sum = scores['test_tp'].sum()
    tn_sum = scores['test_tn'].sum()
    fp_sum = scores['test_fp'].sum()
    fn_sum = scores['test_fn'].sum()
    tp_mean = scores['test_tp'].mean()
    tn_mean = scores['test_tn'].mean()
    fp_mean = scores['test_fp'].mean()
    fn_mean = scores['test_fn'].mean()

    list_to_write_to_file = [counter, n, n_error_class, n_non_error_class, sampling_frequency,
                             ts_fresh_window_length, ts_fresh_window_end, ts_fresh_minimal_features,
                             target_col, target_errorCode, rand_state, sampling_percentage, balance, n_classifier, ensemble_voting,
                             ml_algorithm, cv, Accuracy, Accuracy_mean, Accuracy_std, Precision, Precision_neg, Precision_mean,
                             Precision_mean_neg, Precision_std, Precision_std_neg, Recall, Recall_neg, Recall_mean,
                             Recall_mean_neg, Recall_std, Recall_std_neg, F1_Score, F1_Score_neg, F1_Score_mean,
                             F1_Score_mean_neg, F1_Score_std, F1_Score_std_neg, tp_cv, tn_cv, fp_cv, fn_cv, tp_sum, tn_sum,
                             fp_sum, fn_sum, tp_mean, tn_mean, fp_mean, fn_mean]

    ws1.append(list_to_write_to_file)

    wb.save(filename='../results/Evaluation_results_EUS.xlsx')

    if eval_each_clf:
        for clf in all_scores:
            wb = load_workbook(filename='../results/Evaluation_results_EUS.xlsx')
            ws1 = wb.active
            counter = len(list(ws1.rows))
            n = len(X)
            n_error_class = len(y[y == config['target_errorCode']])
            n_non_error_class = len(y[y != config['target_errorCode']])
            sampling_frequency = config['sampling_frequency']
            ts_fresh_window_length = config['ts_fresh_window_length']
            ts_fresh_window_end = config['ts_fresh_window_end']
            ts_fresh_minimal_features = config['ts_fresh_minimal_features']
            target_col = config['target_col']
            target_errorCode = config['target_errorCode']
            rand_state = str(all_scores[clf]['random_state'])
            sampling_percentage = config['sampling_percentage']
            balance = config['balance_ratio']
            n_classifier = len(all_scores[clf]['random_state'])
            ensemble_voting = config['ensemble_voting']
            ml_algorithm = str([clf_dict[clf]])
            cv = config['cv']
            Accuracy = str(all_scores[clf]['test_accuracy'])
            Accuracy_mean = all_scores[clf]['test_accuracy'].mean()
            Accuracy_std = all_scores[clf]['test_accuracy'].std()
            Precision = str(all_scores[clf]['test_precision'])
            Precision_neg = str(all_scores[clf]['test_precision_neg'])
            Precision_mean = all_scores[clf]['test_precision'].mean()
            Precision_mean_neg = all_scores[clf]['test_precision_neg'].mean()
            Precision_std = all_scores[clf]['test_precision'].std()
            Precision_std_neg = all_scores[clf]['test_precision_neg'].std()
            Recall = str(all_scores[clf]['test_recall'])
            Recall_neg = str(all_scores[clf]['test_recall_neg'])
            Recall_mean = all_scores[clf]['test_recall'].mean()
            Recall_mean_neg = all_scores[clf]['test_recall_neg'].mean()
            Recall_std = all_scores[clf]['test_recall'].std()
            Recall_std_neg = all_scores[clf]['test_recall_neg'].std()
            F1_Score = str(all_scores[clf]['test_f1_score'])
            F1_Score_neg = str(all_scores[clf]['test_f1_score_neg'])
            F1_Score_mean = all_scores[clf]['test_f1_score'].mean()
            F1_Score_mean_neg = all_scores[clf]['test_f1_score_neg'].mean()
            F1_Score_std = all_scores[clf]['test_f1_score'].std()
            F1_Score_std_neg = all_scores[clf]['test_f1_score_neg'].std()
            tp_cv = str(all_scores[clf]['test_tp'])
            tn_cv = str(all_scores[clf]['test_tn'])
            fp_cv = str(all_scores[clf]['test_fp'])
            fn_cv = str(all_scores[clf]['test_fn'])
            tp_sum = all_scores[clf]['test_tp'].sum()
            tn_sum = all_scores[clf]['test_tn'].sum()
            fp_sum = all_scores[clf]['test_fp'].sum()
            fn_sum = all_scores[clf]['test_fn'].sum()
            tp_mean = all_scores[clf]['test_tp'].mean()
            tn_mean = all_scores[clf]['test_tn'].mean()
            fp_mean = all_scores[clf]['test_fp'].mean()
            fn_mean = all_scores[clf]['test_fn'].mean()

            list_to_write_to_file = [counter, n, n_error_class, n_non_error_class, sampling_frequency, ts_fresh_window_length, 
                                     ts_fresh_window_end, ts_fresh_minimal_features, target_col, target_errorCode, 
                                     rand_state, sampling_percentage, balance, n_classifier, ensemble_voting, 
                                     ml_algorithm, cv, Accuracy, Accuracy_mean, Accuracy_std, Precision, Precision_neg, 
                                     Precision_mean, Precision_mean_neg, Precision_std, Precision_std_neg, Recall, Recall_neg, 
                                     Recall_mean, Recall_mean_neg, Recall_std, Recall_std_neg, F1_Score, F1_Score_neg, F1_Score_mean, 
                                     F1_Score_mean_neg, F1_Score_std, F1_Score_std_neg, tp_cv, tn_cv, fp_cv, fn_cv, tp_sum, 
                                     tn_sum, fp_sum, fn_sum, tp_mean, tn_mean, fp_mean, fn_mean]

            ws1.append(list_to_write_to_file)

            wb.save(filename='../results/Evaluation_results_EUS.xlsx')

    return scores
